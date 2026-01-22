"""
Report Generator Module
Generate Excel and CSV reports from package data
"""

import pandas as pd
from typing import Dict, List, Optional
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
import csv
import logging
from datetime import datetime

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ReportGenerator:
    """Generate formatted reports from package data"""
    
    # Column display names (Vietnamese) - includes both mapped and original column names
    COLUMN_NAMES = {
        # Mapped columns
        'source': 'Nguồn',
        'package_code': 'Mã gói',
        'package_name': 'Tên gói',
        'price': 'Giá (VNĐ)',
        'cycle_days': 'Chu kỳ (ngày)',
        'duration': 'Thời gian hiệu lực',
        'data_gb': 'Data (GB)',
        'data_limit_behavior': 'Hết data',
        'voice_minutes': 'Phút gọi',
        'sms_count': 'SMS',
        'package_type': 'Loại gói',
        'description': 'Mô tả',
        'full_description': 'Chi tiết',
        'registration_syntax': 'Cú pháp ĐK',
        'cancellation_syntax': 'Cú pháp hủy',
        'check_syntax': 'Cú pháp tra cứu',
        'eligibility': 'Điều kiện',
        'renewal_policy': 'Chính sách GH',
        'support_hotline': 'Hotline',
        'original_link': 'Link gốc',
        'benefits': 'Lợi ích',
        'variants': 'Biến thể',
        'related_packages': 'Gói liên quan',
        'benefit_free_internal_calls': 'Gọi nội mạng',
        'benefit_free_external_calls': 'Gọi ngoại mạng',
        'benefit_free_sms': 'SMS miễn phí',
        'benefit_free_social_media_data': 'Data MXH',
        'benefit_free_tv': 'TV miễn phí',
        'benefit_other_benefits': 'Lợi ích khác',
        'source_file': 'File nguồn',
        'relationship_type': 'Loại quan hệ',
        # Original column names from CSV
        'code': 'Mã gói (gốc)',
        'full_name': 'Tên đầy đủ',
        'cycle': 'Chu kỳ',
        'data_size': 'Dung lượng data',
        'source_url': 'URL nguồn',
        'registration': 'Đăng ký',
        'notes': 'Ghi chú'
    }
    
    def __init__(self):
        """Initialize report generator"""
        pass
    
    def format_package_info(self, package: Dict) -> Dict:
        """
        Format package information for display
        
        Args:
            package: Package dictionary
            
        Returns:
            Formatted package dictionary
        """
        formatted = package.copy()
        
        # Remove internal fields
        for key in ['_similarity_score', '_match_field', '_search_string']:
            formatted.pop(key, None)
        
        # Format price
        if 'price' in formatted and pd.notna(formatted['price']):
            try:
                formatted['price'] = f"{float(formatted['price']):,.0f}"
            except (ValueError, TypeError):
                pass
        
        # Format data
        if 'data_gb' in formatted and pd.notna(formatted['data_gb']):
            try:
                formatted['data_gb'] = f"{float(formatted['data_gb']):.2f}"
            except (ValueError, TypeError):
                pass
        
        # Format cycle
        if 'cycle_days' in formatted and pd.notna(formatted['cycle_days']):
            try:
                days = float(formatted['cycle_days'])
                if days >= 30:
                    months = days / 30
                    formatted['cycle_days'] = f"{months:.0f} tháng"
                else:
                    formatted['cycle_days'] = f"{days:.0f} ngày"
            except (ValueError, TypeError):
                pass
        
        return formatted
    
    def generate_csv(
        self,
        packages: List[Dict],
        output_path: str,
        include_internal: bool = False
    ) -> str:
        """
        Generate CSV report
        
        Args:
            packages: List of package dictionaries
            output_path: Output file path
            include_internal: Include internal fields like similarity score
            
        Returns:
            Path to generated file
        """
        if not packages:
            raise ValueError("No packages to export")
        
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Clean packages
        cleaned_packages = []
        for pkg in packages:
            cleaned = pkg.copy()
            if not include_internal:
                for key in ['_similarity_score', '_match_field', '_search_string']:
                    cleaned.pop(key, None)
            cleaned_packages.append(cleaned)
        
        # Convert to DataFrame and save
        df = pd.DataFrame(cleaned_packages)
        df.to_csv(output_path, index=False, encoding='utf-8-sig')
        
        logger.info(f"✅ CSV report generated: {output_path}")
        return str(output_path)
    
    def generate_excel(
        self,
        packages: List[Dict],
        output_path: str,
        include_similarity: bool = False
    ) -> str:
        """
        Generate formatted Excel report
        
        Args:
            packages: List of package dictionaries
            output_path: Output file path
            include_similarity: Include similarity scores in output
            
        Returns:
            Path to generated file
        """
        if not packages:
            raise ValueError("No packages to export")
        
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Package Report"
        
        # Clean packages
        cleaned_packages = []
        for pkg in packages:
            cleaned = pkg.copy()
            if not include_similarity:
                for key in ['_similarity_score', '_match_field', '_search_string']:
                    cleaned.pop(key, None)
            cleaned_packages.append(cleaned)
        
        # Convert to DataFrame
        df = pd.DataFrame(cleaned_packages)
        
        # Add similarity score column if requested
        if include_similarity and '_similarity_score' in packages[0]:
            df['_similarity_score'] = [p.get('_similarity_score', 0) for p in packages]
            self.COLUMN_NAMES['_similarity_score'] = 'Độ tương đồng (%)'
        
        # Write headers with Vietnamese names
        headers = []
        for col in df.columns:
            header_name = self.COLUMN_NAMES.get(col, col)
            headers.append(header_name)
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            self._style_header_cell(cell)
        
        # Write data
        for row_idx, record in enumerate(df.to_dict('records'), 2):
            for col_idx, col in enumerate(df.columns, 1):
                value = record.get(col, '')
                
                # Handle NaN and None
                if pd.isna(value) or value is None:
                    value = ''
                
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                self._style_data_cell(cell, col)
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws, df)
        
        # Freeze first row
        ws.freeze_panes = 'A2'
        
        # Add filters
        ws.auto_filter.ref = ws.dimensions
        
        # Add metadata sheet
        self._add_metadata_sheet(wb, packages)
        
        # Save workbook
        wb.save(output_path)
        
        logger.info(f"✅ Excel report generated: {output_path}")
        return str(output_path)
    
    def _style_header_cell(self, cell):
        """Apply styling to header cell"""
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def _style_data_cell(self, cell, column_name: str):
        """Apply styling to data cell"""
        cell.border = Border(
            left=Side(style='thin', color="E0E0E0"),
            right=Side(style='thin', color="E0E0E0"),
            top=Side(style='thin', color="E0E0E0"),
            bottom=Side(style='thin', color="E0E0E0")
        )
        
        # Special formatting for specific columns
        if column_name == 'price':
            cell.alignment = Alignment(horizontal="right")
            # Try to format as number
            try:
                if cell.value and str(cell.value).strip():
                    cell.value = float(cell.value)
                    cell.number_format = '#,##0'
            except (ValueError, TypeError):
                pass
        
        elif column_name in ['data_gb', 'cycle_days', 'sms_count', 'voice_minutes']:
            cell.alignment = Alignment(horizontal="right")
            try:
                if cell.value and str(cell.value).strip():
                    cell.value = float(cell.value)
                    cell.number_format = '0.00' if column_name == 'data_gb' else '0'
            except (ValueError, TypeError):
                pass
        
        elif column_name == 'package_code':
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        elif column_name in ['description', 'full_description']:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        else:
            cell.alignment = Alignment(vertical="top")
    
    def _auto_adjust_columns(self, ws: Worksheet, df: pd.DataFrame):
        """Auto-adjust column widths based on content"""
        for col_idx, col in enumerate(df.columns, 1):
            column_letter = get_column_letter(col_idx)
            
            # Default widths for specific columns
            if col in ['description', 'full_description']:
                ws.column_dimensions[column_letter].width = 50
            elif col in ['eligibility', 'renewal_policy']:
                ws.column_dimensions[column_letter].width = 40
            elif col in ['registration_syntax', 'cancellation_syntax', 'check_syntax']:
                ws.column_dimensions[column_letter].width = 30
            elif col == 'original_link':
                ws.column_dimensions[column_letter].width = 35
            elif col in ['package_code', 'package_name']:
                ws.column_dimensions[column_letter].width = 15
            elif col == 'source':
                ws.column_dimensions[column_letter].width = 12
            else:
                # Auto-size based on content
                max_length = 0
                column = ws[column_letter]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = max(adjusted_width, 10)
    
    def _add_metadata_sheet(self, wb: Workbook, packages: List[Dict]):
        """Add metadata sheet with report info"""
        ws = wb.create_sheet("Thông tin")
        
        # Report metadata
        metadata = [
            ["Báo cáo gói cước viễn thông", ""],
            ["Ngày tạo", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Số lượng gói", len(packages)],
            ["", ""],
            ["Nguồn dữ liệu", ""],
        ]
        
        # Count by source
        sources = {}
        for pkg in packages:
            source = pkg.get('source', 'unknown')
            sources[source] = sources.get(source, 0) + 1
        
        for source, count in sources.items():
            metadata.append([f"  - {source}", count])
        
        # Write metadata
        for row_idx, (key, value) in enumerate(metadata, 1):
            ws.cell(row=row_idx, column=1, value=key).font = Font(bold=True)
            ws.cell(row=row_idx, column=2, value=value)
        
        # Style
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30
    
    def generate_summary_report(
        self,
        packages: List[Dict],
        output_path: str
    ) -> str:
        """
        Generate summary report with statistics
        
        Args:
            packages: List of package dictionaries
            output_path: Output file path
            
        Returns:
            Path to generated file
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Convert to DataFrame for analysis
        df = pd.DataFrame(packages)
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write("=" * 60 + "\n")
            f.write("BÁO CÁO TỔNG HỢP GÓI CƯỚC\n")
            f.write("=" * 60 + "\n\n")
            
            f.write(f"Ngày tạo: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Tổng số gói: {len(packages)}\n\n")
            
            # By source
            f.write("Phân bổ theo nguồn:\n")
            f.write("-" * 40 + "\n")
            source_counts = df['source'].value_counts()
            for source, count in source_counts.items():
                f.write(f"  {source}: {count} gói ({count/len(df)*100:.1f}%)\n")
            f.write("\n")
            
            # Price statistics
            f.write("Thống kê giá:\n")
            f.write("-" * 40 + "\n")
            f.write(f"  Min: {df['price'].min():,.0f} VNĐ\n")
            f.write(f"  Max: {df['price'].max():,.0f} VNĐ\n")
            f.write(f"  Trung bình: {df['price'].mean():,.0f} VNĐ\n")
            f.write(f"  Trung vị: {df['price'].median():,.0f} VNĐ\n")
            f.write("\n")
            
            # Data statistics
            if 'data_gb' in df.columns and df['data_gb'].notna().any():
                f.write("Thống kê dung lượng:\n")
                f.write("-" * 40 + "\n")
                f.write(f"  Min: {df['data_gb'].min():.2f} GB\n")
                f.write(f"  Max: {df['data_gb'].max():.2f} GB\n")
                f.write(f"  Trung bình: {df['data_gb'].mean():.2f} GB\n")
                f.write("\n")
            
            # Top packages by price
            f.write("Top 10 gói đắt nhất:\n")
            f.write("-" * 40 + "\n")
            top_expensive = df.nlargest(10, 'price')
            for idx, row in enumerate(top_expensive.itertuples(), 1):
                f.write(f"  {idx}. {row.package_code}: {row.price:,.0f} VNĐ\n")
            f.write("\n")
            
            # Top packages by data
            if 'data_gb' in df.columns and df['data_gb'].notna().any():
                f.write("Top 10 gói data lớn nhất:\n")
                f.write("-" * 40 + "\n")
                top_data = df.nlargest(10, 'data_gb')
                for idx, row in enumerate(top_data.itertuples(), 1):
                    f.write(f"  {idx}. {row.package_code}: {row.data_gb:.2f} GB\n")
        
        logger.info(f"✅ Summary report generated: {output_path}")
        return str(output_path)


if __name__ == "__main__":
    # Test report generator
    from data_loader import PackageDataLoader
    
    try:
        loader = PackageDataLoader()
        df = loader.load_data()
        packages = df.head(20).to_dict('records')
        
        generator = ReportGenerator()
        
        # Test CSV export
        csv_path = generator.generate_csv(packages, "test_report.csv")
        print(f"✅ CSV generated: {csv_path}")
        
        # Test Excel export
        excel_path = generator.generate_excel(packages, "test_report.xlsx")
        print(f"✅ Excel generated: {excel_path}")
        
        # Test summary
        summary_path = generator.generate_summary_report(packages, "test_summary.txt")
        print(f"✅ Summary generated: {summary_path}")
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
